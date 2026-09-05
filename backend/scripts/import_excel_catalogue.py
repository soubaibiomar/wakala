"""Import the Wakala Excel catalogue into the live showroom table.

Usage (from the backend directory):
    python scripts/import_excel_catalogue.py path/to/wakala-catalogue.xlsx

The importer is deliberately idempotent: the same workbook can be imported
again without creating duplicate showroom vehicles.
"""

from __future__ import annotations

import asyncio
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, update

from app.core.database import async_session_factory
from app.core.security import hash_password
from app.models.listing import Listing
from app.models.user import User
from app.models.vehicle import Vehicle
from app.models.catalog import BrandCatalog, ModelCatalog, PowertrainCatalog, TrimCatalog


SOURCE = "wakala_excel"
DEFAULT_CITY = "Casablanca"
FALLBACK_IMAGE = "https://images.unsplash.com/photo-1549399542-7e3f8b79c341?q=80&w=1200"


def clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def number(value: object, default: int | None = None) -> int | None:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return int(round(value))
    match = re.search(r"-?\d[\d\s.,]*", clean(value))
    if not match:
        return default
    raw = re.sub(r"[^0-9-]", "", match.group(0))
    try:
        return int(raw)
    except ValueError:
        return default


def fuel_type(value: str) -> str:
    value = value.lower()
    if "recharge" in value or "phev" in value:
        return "hybride_rechargeable"
    if "élect" in value or "elect" in value:
        return "electrique"
    if "hybrid" in value:
        return "hybride"
    if "gpl" in value:
        return "gpl"
    if "hydrog" in value:
        return "hydrogene"
    if "diesel" in value:
        return "diesel"
    return "essence"


def body_type(model: str, finish: str, brand: str = "", row: dict[str, object] | None = None) -> str:
    if row:
        for k in ("[AG] Carrosserie", "Carrosserie", "carrosserie", "Type Carrosserie"):
            val = clean(row.get(k)).lower()
            if val in ("citadine", "berline", "suv", "break", "coupe", "cabriolet", "monospace", "utilitaire", "pick_up"):
                return val

    # Fallback: try data_pipeline catalogue_mapping if available
    try:
        from data_pipeline.scripts.catalogue_mapping import infer_body_type
        return infer_body_type(brand, model, finish, None, False, None, row=row)
    except Exception:
        pass

    text = f"{brand} {model} {finish}".lower()
    if any(x in text for x in ("pick-up", "pickup", "pik-up", "hilux", "ranger", "d-max", "l200", "navara", "landtrek", "titano")):
        return "pick_up"
    if any(x in text for x in ("monospace", "space", "staria", "carnival", "spacetourer", "zafira")):
        return "monospace"
    if any(x in text for x in ("break", "touring", "estate", "sw", "wagon", "avant")):
        return "break"
    if any(x in text for x in ("coupé", "coupe", "a110", "gtb", "911")):
        return "coupe"
    if any(x in text for x in ("cabrio", "convertible", "spider", "spyder", "roadster")):
        return "cabriolet"
    if any(x in text for x in ("utilitaire", "van", "fourgon", "transit", "berlingo", "partner", "kangoo", "caddy", "combo", "express")):
        return "utilitaire"
    if any(x in text for x in ("citadine", "city", "mini", "clio", "208", "c3", "sandero", "yaris", "polo", "i10", "i20", "picanto", "swift", "micra", "fiesta", "fabia", "ibiza", "corsa", "spring", "500")):
        return "citadine"
    if any(x in text for x in ("berline", "sedan", "saloon", "corolla", "golf", "megane", "octavia", "astra", "classe a", "classe c", "classe e", "serie 3", "serie 5", "a3", "a4", "a6", "passat", "mondeo")):
        return "berline"
    if any(x in text for x in ("suv", "crossover", "4x4", "duster", "tiguan", "tucson", "sportage", "qashqai", "3008", "2008", "5008", "rav4", "kuga", "captur", "austral", "t-roc", "t-cross", "x1", "x3", "x5", "q3", "q5", "gla", "glc")):
        return "suv"
    return "berline"



def transmission(value: str) -> str:
    value = value.lower()
    if any(x in value for x in ("bvm", "manuel", "manual")):
        return "manuelle"
    return "automatique"


def parse_price(value: object) -> float | None:
    result = number(value)
    if result is None or result <= 0:
        return None
    # The delivered workbook has prices rounded to a clean hundred.
    return float(round(result / 100) * 100)


def ncap_stars(value: object) -> int | None:
    match = re.search(r"([0-5])\s*[★*]", clean(value))
    return int(match.group(1)) if match else None


def slug(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value, flags=re.IGNORECASE)
    return value.strip("-") or "vehicle"


def row_to_vehicle(row: dict[str, object]) -> dict[str, object] | None:
    brand = clean(row.get("[A] Marque"))
    model = clean(row.get("[E] Modèle"))
    version = clean(row.get("[F] Finition / Variante"))
    price = parse_price(row.get("[I] Prix (DH)"))
    if not brand or not model or price is None:
        return None

    ncap = clean(row.get("[R] Sécurité NCAP"))
    source_url = clean(row.get("[H] Fiche Finition Officielle"))
    bt = body_type(model, version, brand=brand, row=row)
    return {
        "brand": brand[:100],
        "model": model[:100],
        "version": version[:200] or None,
        "year": 2026,
        "mileage": 0,
        "fuel_type": fuel_type(clean(row.get("[O] Type Moteur"))),
        "body_type": bt,
        "transmission": transmission(clean(row.get("[P] Transmission"))),
        "engine_power_hp": number(row.get("[N] Puissance (ch)")),
        "color": None,
        "doors": 5,
        "seats": 7 if bt == "monospace" else 5,
        "trunk_volume_l": number(row.get("[J] Coffre (L)")),
        "ncap_rating": ncap[:50] or None,
        "co2_emissions": number(row.get("[Q] CO2 (g/km)")),
        "electric_range_km": number(row.get("[V] Autonomie (km)")),
        "is_4x4": "4x4" in clean(row.get("[P] Transmission")).lower(),
        "condition": "new",
        "source": SOURCE,
        "status": "available",
        "source_url": source_url[:500] or None,
        "city": DEFAULT_CITY,
        "price": price,
        "description": f"{brand} {model} — {version}"[:1000],
        "image_urls": [FALLBACK_IMAGE],
    }


async def import_new_car_catalog_row(db, row: dict[str, object]) -> bool:
    """Synchronize the model/trim showroom used by /v1/new-cars as well."""
    brand_name = clean(row.get("[A] Marque"))
    model_name = clean(row.get("[E] Modèle"))
    trim_name = clean(row.get("[F] Finition / Variante"))
    price = parse_price(row.get("[I] Prix (DH)"))
    if not brand_name or not model_name or not trim_name or price is None:
        return False

    brand_result = await db.execute(select(BrandCatalog).where(BrandCatalog.name == brand_name))
    brand = brand_result.scalars().first()
    if brand is None:
        brand = BrandCatalog(name=brand_name[:100], slug=slug(brand_name), is_active=True)
        db.add(brand)
        await db.flush()

    model_result = await db.execute(
        select(ModelCatalog).where(ModelCatalog.brand_id == brand.id, ModelCatalog.name == model_name)
    )
    model = model_result.scalars().first()
    model_body = body_type(model_name, trim_name, brand=brand_name, row=row)
    if model is None:
        model = ModelCatalog(
            brand_id=brand.id,
            name=model_name[:100],
            slug=slug(model_name),
            body_type=model_body,
            year_start=2026,
            hero_image_url=None,
        )
        db.add(model)
        await db.flush()
    else:
        model.body_type = model_body

    fuel = fuel_type(clean(row.get("[O] Type Moteur"))).upper()
    gearbox = transmission(clean(row.get("[P] Transmission"))).upper()
    power = number(row.get("[N] Puissance (ch)"))
    pt_result = await db.execute(
        select(PowertrainCatalog).where(
            PowertrainCatalog.model_id == model.id,
            PowertrainCatalog.name == trim_name[:150],
        )
    )
    powertrain = pt_result.scalars().first()
    if powertrain is None:
        powertrain = PowertrainCatalog(
            model_id=model.id,
            name=trim_name[:150],
            fuel_type=fuel,
            fiscal_power_cv=6,
            engine_power_hp=power,
            transmission=gearbox,
            drivetrain="AWD" if "4x4" in clean(row.get("[P] Transmission")).lower() else "FWD",
        )
        db.add(powertrain)
        await db.flush()
    else:
        powertrain.fuel_type = fuel
        powertrain.engine_power_hp = power
        powertrain.transmission = gearbox

    trim_result = await db.execute(
        select(TrimCatalog).where(
            TrimCatalog.model_id == model.id,
            TrimCatalog.name == trim_name,
        )
    )
    trim = trim_result.scalars().first()
    if trim is None:
        trim = TrimCatalog(
            model_id=model.id,
            powertrain_id=powertrain.id,
            name=trim_name[:100],
            slug=slug(trim_name),
        )
        db.add(trim)
    trim.powertrain_id = powertrain.id
    trim.price_new_mad = price
    trim.promo_price_mad = None
    trim.is_promo = False
    trim.trunk_capacity_l = number(row.get("[J] Coffre (L)"))
    trim.euro_ncap_stars = ncap_stars(row.get("[R] Sécurité NCAP"))
    trim.is_available_in_morocco = True
    await db.flush()
    return True


async def ensure_import_user(db) -> uuid.UUID:
    result = await db.execute(select(User).where(User.email == "excel_catalogue@wakala.ma"))
    user = result.scalar_one_or_none()
    if user:
        return user.id
    user = User(
        id=uuid.uuid4(),
        email="excel_catalogue@wakala.ma",
        hashed_password=hash_password(uuid.uuid4().hex),
        full_name="Excel Catalogue Import",
        phone="+212600000002",
        role="seller",
        is_verified=True,
        is_pro=True,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db.add(user)
    await db.flush()
    return user.id


async def import_workbook(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in workbook.worksheets if "Catalogue" in s.title), workbook.worksheets[0])
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    imported: list[dict[str, object]] = []
    source_rows: list[dict[str, object]] = []
    for values in rows:
        row = dict(zip(headers, values))
        source_rows.append(row)
        vehicle = row_to_vehicle(row)
        if vehicle:
            imported.append(vehicle)
    if not imported:
        raise RuntimeError("No valid vehicle rows found in the workbook")

    async with async_session_factory() as db:
        seller_id = await ensure_import_user(db)
        # The Excel workbook is the source of truth for the new-car showroom.
        # Keep historical seed rows recoverable, but hide them from the live
        # API before reactivating/updating the exact workbook rows below.
        await db.execute(
            update(TrimCatalog)
            .values(is_available_in_morocco=False)
        )
        new_catalog_count = 0
        for row in source_rows:
            if await import_new_car_catalog_row(db, row):
                new_catalog_count += 1
        # Hide previous catalogue snapshots before replacing them. This keeps
        # the showroom count equal to the workbook instead of appending the
        # Excel rows to the old demo catalogue. User listings from other
        # sources are left untouched.
        await db.execute(
            update(Vehicle)
            .where(Vehicle.source.in_((SOURCE, "wakala_catalogue")))
            .values(status="deleted", updated_at=datetime.now(timezone.utc))
        )

        now = datetime.now(timezone.utc)
        for item in imported:
            existing_result = await db.execute(
                select(Vehicle).where(
                    Vehicle.source == SOURCE,
                    Vehicle.brand == item["brand"],
                    Vehicle.model == item["model"],
                    Vehicle.version == item["version"],
                )
            )
            vehicle = existing_result.scalar_one_or_none()
            if vehicle is None:
                vehicle = Vehicle(id=uuid.uuid4(), seller_id=seller_id)
                db.add(vehicle)
            for key, value in item.items():
                if key != "image_urls":
                    setattr(vehicle, key, value)
            vehicle.seller_id = seller_id
            vehicle.created_at = now
            vehicle.updated_at = now
            await db.flush()

            listing_result = await db.execute(select(Listing).where(Listing.vehicle_id == vehicle.id))
            listing = listing_result.scalar_one_or_none()
            if listing is None:
                listing = Listing(id=uuid.uuid4(), vehicle_id=vehicle.id)
                db.add(listing)
            listing.status = "active"
            listing.published_at = now
            listing.images_urls = item["image_urls"]
            listing.updated_at = now

        await db.commit()
    print(f"Synchronized {new_catalog_count} new-car trims")
    return len(imported)


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python scripts/import_excel_catalogue.py <workbook.xlsx>")
    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.is_file():
        raise SystemExit(f"Workbook not found: {path}")
    count = asyncio.run(import_workbook(path))
    print(f"Imported {count} vehicles from {path}")


if __name__ == "__main__":
    main()
