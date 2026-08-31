#!/usr/bin/env python3
"""
update_excel_catalogue_with_official_data.py — Met à jour D:/Projet automobile/wakala-catalogue.xlsx
avec les colonnes Couleurs Officielles & HEX, Options & Packs d'Équipements, et Sites Officiels Constructeurs.
"""

import os
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from data_pipeline.scripts.official_brand_catalog_data import (
    OFFICIAL_BRAND_COLORS,
    DEFAULT_UNIVERSAL_COLORS,
    get_official_colors_for_vehicle,
    get_official_options_for_vehicle,
)

OFFICIAL_WEBSITES = {
    'dacia': 'https://www.dacia.ma',
    'renault': 'https://www.renault.ma',
    'peugeot': 'https://www.peugeot.ma',
    'citroen': 'https://www.citroen.ma',
    'citroën': 'https://www.citroen.ma',
    'volkswagen': 'https://www.volkswagen.ma',
    'audi': 'https://www.audi.ma',
    'skoda': 'https://www.skoda.ma',
    'seat': 'https://www.seat.ma',
    'cupra': 'https://www.cupraofficial.ma',
    'bmw': 'https://www.bmw.ma',
    'mini': 'https://www.mini.ma',
    'mercedes': 'https://www.mercedes-benz.ma',
    'mercedes-benz': 'https://www.mercedes-benz.ma',
    'porsche': 'https://www.porsche.com/middle-east/_morocco_/',
    'toyota': 'https://www.toyota.ma',
    'lexus': 'https://www.lexus.ma',
    'hyundai': 'https://www.hyundai.ma',
    'kia': 'https://www.kia.ma',
    'nissan': 'https://www.nissan.ma',
    'ford': 'https://www.ford.ma',
    'fiat': 'https://www.fiat.ma',
    'alfa': 'https://www.alfaromeo.ma',
    'alfa romeo': 'https://www.alfaromeo.ma',
    'jeep': 'https://www.jeep.ma',
    'land rover': 'https://www.landrover-morocco.com',
    'jaguar': 'https://www.jaguar-morocco.com',
    'volvo': 'https://www.volvocars.com/fr-ma',
    'suzuki': 'https://www.suzuki.ma',
    'honda': 'https://www.honda.ma',
    'mitsubishi': 'https://www.mitsubishi-motors.ma',
    'mg': 'https://www.mgmotor.ma',
    'byd': 'https://www.byd.ma',
    'chery': 'https://www.chery.ma',
    'geely': 'https://www.geely.ma',
    'gwm': 'https://www.gwm.ma',
    'haval': 'https://www.gwm.ma',
    'changan': 'https://www.changan.ma',
    'dfsk': 'https://www.dfsk.ma',
    'omoda': 'https://www.omoda.ma',
    'seres': 'https://www.seres.ma',
    'xpeng': 'https://www.xpeng.com',
    'zeekr': 'https://www.zeekr.com',
    'leapmotor': 'https://www.leapmotor.com',
    'baic': 'https://www.baic.ma',
    'bentley': 'https://www.bentleymotors.com',
    'ds': 'https://www.dsautomobiles.ma',
    'opel': 'https://www.opel.ma',
}


def get_website(brand_str: str) -> str:
    b_clean = str(brand_str or '').lower().strip()
    for k, v in OFFICIAL_WEBSITES.items():
        if k in b_clean:
            return v
    return f'https://www.{b_clean.replace(" ", "")}.ma'


def main():
    file_path = Path("D:/Projet automobile/wakala-catalogue.xlsx")
    if not file_path.exists():
        file_path = PROJECT_ROOT.parent / "wakala-catalogue.xlsx"

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Catalogue']

    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    data_font = Font(name='Segoe UI', size=9)
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # 1. Update Column Headers in Catalogue Sheet
    ws.cell(row=3, column=25, value='Couleurs Officielles & HEX')
    ws.cell(row=3, column=26, value='Options & Packs Équipements')
    ws.cell(row=3, column=27, value='Site Web Officiel Marque')

    for c in range(25, 28):
        cell = ws.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Fill rows in Catalogue sheet
    for r in range(4, ws.max_row + 1):
        brand_val = str(ws.cell(row=r, column=1).value or '').strip()
        model_val = str(ws.cell(row=r, column=2).value or '').strip()
        price_val = ws.cell(row=r, column=4).value or 150000

        if not brand_val or not model_val:
            continue

        # Colors string
        colors = get_official_colors_for_vehicle(brand_val, model_val)
        colors_parts = []
        for c in colors:
            price_tag = "0 DH" if c.get("price_mad", 0) == 0 else f"+{c.get('price_mad')} DH"
            colors_parts.append(f"{c['name']} ({c['hex']} · {price_tag})")
        colors_str = " | ".join(colors_parts)

        # Options string
        body_hint = 'suv' if any(kw in model_val.lower() for kw in ['duster', 'tucson', 'sportage', 'rav4', 'tiguan', 'qashqai', 'suv']) else 'berline'
        opts = get_official_options_for_vehicle(brand_val, model_val, body_hint, float(price_val or 150000))
        opts_parts = []
        for o in opts[:6]:
            delta_tag = "Série" if o.get("price_delta", 0) == 0 else f"+{o.get('price_delta')} DH"
            opts_parts.append(f"{o['category'].upper()}: {o['name']} ({delta_tag})")
        opts_str = " | ".join(opts_parts)

        # Website
        site_url = get_website(brand_val)

        c25 = ws.cell(row=r, column=25, value=colors_str)
        c26 = ws.cell(row=r, column=26, value=opts_str)
        c27 = ws.cell(row=r, column=27, value=site_url)

        for cell in (c25, c26, c27):
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = Alignment(vertical='center')

    # 2. Add / Refresh dedicated Sheet: 'Nuanciers & Options Marques'
    sheet_name = 'Nuanciers & Options Marques'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws_no = wb.create_sheet(title=sheet_name)
    ws_no.views.sheetView[0].showGridLines = True

    ws_no.cell(row=1, column=1, value='WAKALA — Répertoire Officiel des Nuanciers, Couleurs HEX et Options Constructeurs')
    ws_no.cell(row=1, column=1).font = Font(name='Segoe UI', size=14, bold=True, color='1E3A8A')

    headers_no = ['Marque', 'Site Web Officiel', 'Couleur Nom', 'Code HEX', 'Finition / Type', 'Surcoût DH', 'Option / Pack Type', 'Nom Équipement', 'Tarif Option DH']
    for col_idx, h in enumerate(headers_no, start=1):
        cell = ws_no.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    curr_row = 4
    for b_key, b_colors in OFFICIAL_BRAND_COLORS.items():
        b_display = b_key.capitalize()
        if b_key == 'byd': b_display = 'BYD'
        elif b_key == 'bmw': b_display = 'BMW'
        elif b_key == 'mg': b_display = 'MG'
        elif b_key == 'mercedes': b_display = 'Mercedes-Benz'
        elif b_key == 'dacia': b_display = 'Dacia'
        elif b_key == 'renault': b_display = 'Renault'
        elif b_key == 'peugeot': b_display = 'Peugeot'
        elif b_key == 'volkswagen': b_display = 'Volkswagen'

        b_url = get_website(b_display)
        b_opts = get_official_options_for_vehicle(b_display, 'Standard', 'suv' if 'dacia' in b_key or 'jeep' in b_key else 'berline', 200000)

        max_len = max(len(b_colors), len(b_opts))
        for i in range(max_len):
            c_item = b_colors[i] if i < len(b_colors) else None
            o_item = b_opts[i] if i < len(b_opts) else None

            ws_no.cell(row=curr_row, column=1, value=b_display if i == 0 else '').font = data_font
            ws_no.cell(row=curr_row, column=2, value=b_url if i == 0 else '').font = data_font

            if c_item:
                ws_no.cell(row=curr_row, column=3, value=c_item['name']).font = data_font
                ws_no.cell(row=curr_row, column=4, value=c_item['hex']).font = data_font
                ws_no.cell(row=curr_row, column=5, value='Série / Opaque' if c_item.get('price_mad', 0) == 0 else 'Métallisé / Nacré').font = data_font
                ws_no.cell(row=curr_row, column=6, value=c_item.get('price_mad', 0)).font = data_font

            if o_item:
                ws_no.cell(row=curr_row, column=7, value=o_item['category'].upper()).font = data_font
                ws_no.cell(row=curr_row, column=8, value=o_item['name']).font = data_font
                ws_no.cell(row=curr_row, column=9, value=o_item.get('price_delta', 0)).font = data_font

            for c in range(1, 10):
                ws_no.cell(row=curr_row, column=c).border = border_thin

            curr_row += 1

    wb.save(file_path)
    print(f"✅ Fichier Excel mis à jour avec succès : {file_path}")


if __name__ == "__main__":
    main()
