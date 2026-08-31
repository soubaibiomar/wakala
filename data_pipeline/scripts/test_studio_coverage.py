import json
import re
from pathlib import Path
import openpyxl

def main():
    ts_file = Path("d:/Projet automobile/vente-auto-platform/frontend/src/utils/vehicleImageCatalogData.ts")
    with open(ts_file, "r", encoding="utf-8") as f:
        content = f.read()

    studio_map = {}
    current_brand = None
    for line in content.splitlines():
        line = line.strip()
        brand_match = re.match(r'^"([^"]+)":\s*\{', line)
        if brand_match:
            current_brand = brand_match.group(1).lower().strip()
            studio_map[current_brand] = {}
            continue
        item_match = re.match(r'^"([^"]+)":\s*"([^"]+)"', line)
        if item_match and current_brand:
            k = item_match.group(1).lower().strip()
            v = item_match.group(2).strip()
            studio_map[current_brand][k] = v

    excel_file = Path("d:/Projet automobile/wakala-catalogue.xlsx")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb["Catalogue Véhicules"] if "Catalogue Véhicules" in wb.sheetnames else wb.worksheets[0]

    matched = 0
    total_trims = 0
    unmatched_models = set()

    for r in range(2, sheet.max_row + 1):
        brand = sheet.cell(r, 1).value
        model = sheet.cell(r, 5).value
        if not brand or not model:
            continue

        b_key = str(brand).lower().strip()
        m_key = str(model).lower().strip()

        total_trims += 1

        img = None
        if b_key in studio_map:
            img = studio_map[b_key].get(m_key)
            if not img:
                for k, v in studio_map[b_key].items():
                    if m_key in k or k in m_key:
                        img = v
                        break
        if img:
            matched += 1
        else:
            unmatched_models.add((brand, model))

    print(f"Total Trims in Excel: {total_trims}")
    print(f"Direct Studio Matched: {matched} ({matched/total_trims*100:.1f}%)")
    print(f"Unmatched distinct models count: {len(unmatched_models)}")
    print("Sample unmatched models:")
    for b, m in sorted(list(unmatched_models))[:20]:
        print(f"  - {b}: {m}")

if __name__ == "__main__":
    main()
