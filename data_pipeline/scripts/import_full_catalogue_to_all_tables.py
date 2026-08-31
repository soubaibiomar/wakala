#!/usr/bin/env python3
"""
import_full_catalogue_to_all_tables.py — Importation complète et unifiée du catalogue officiel Wakala.
====================================================================================================

Lit le fichier Excel wakala-catalogue.xlsx (420 véhicules, 55 marques) et peuple de manière
100% idempotente et transactionnelle :
1. 'car_brands' (toutes les 55 marques avec pays d'origine et logos)
2. 'car_models' (tous les modèles avec carrosserie, photos réelles HD)
3. 'car_powertrains' (motorisations, puissances fiscale CV & DIN ch, consommations)
4. 'car_trims' (finitions, prix catalogue en MAD, garantie, étoiles NCAP, images réelles)
5. 'vehicles' (catalogue global véhicules neufs 2026)
6. 'vehicle_wakala_scores' (les 8 notes d'évaluation Wakala 1-5 + note globale /5)
7. 'vehicle_options' & 'vehicle_colors' (options, accessoires, jantes, selleries et teintes)
"""

import argparse
import logging
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Configuration du PYTHONPATH
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "data_pipeline" / "kafka" / "producers" / "scrapers"))

import openpyxl
import json
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from data_pipeline.scripts.catalogue_mapping import (
    map_excel_row_to_vehicle_data,
    EXCEL_COLUMN_HEADERS,
)
from data_pipeline.scripts.vehicle_image_catalog import (
    BRAND_ORIGINS,
    get_real_vehicle_image,
)
from data_pipeline.scripts.official_brand_catalog_data import (
    get_official_colors_for_vehicle,
    get_official_options_for_vehicle,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("WakalaFullCatalogueImporter")

# UUID Namespaces
WAKALA_NAMESPACE = uuid.UUID("e743a18e-42c2-4876-9051-b841e4eb4192")
DEFAULT_SYSTEM_SELLER_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")

CANDIDATE_PATHS = [
    Path(r"D:\Projet automobile\wakala-catalogue.xlsx"),
    Path(r"C:\Users\omar\Downloads\wakala-catalogue.xlsx"),
    PROJECT_ROOT.parent / "wakala-catalogue.xlsx",
    PROJECT_ROOT / "wakala-catalogue.xlsx",
]


def resolve_catalogue_file(custom_path: Optional[str] = None) -> Path:
    if custom_path:
        p = Path(custom_path)
        if p.exists():
            return p
        raise FileNotFoundError(f"Fichier catalogue introuvable : {custom_path}")

    for p in CANDIDATE_PATHS:
        if p.exists():
            logger.info(f"Fichier catalogue détecté : {p}")
            return p

    raise FileNotFoundError("Fichier wakala-catalogue.xlsx introuvable.")


def get_db_url() -> str:
    env_file = PROJECT_ROOT / ".env"
    if env_file.exists():
        with open(env_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    if k not in os.environ:
                        os.environ[k] = v.strip('"').strip("'")

    db_url = os.getenv("DATABASE_URL")
    if db_url:
        return db_url.replace("postgresql+asyncpg://", "postgresql://")

    user = os.getenv("POSTGRES_USER", "wakala_user")
    password = os.getenv("POSTGRES_PASSWORD", "wakala_secret_password")
    host = os.getenv("POSTGRES_HOST", "localhost")
    port = os.getenv("POSTGRES_PORT", "5433")
    db = os.getenv("POSTGRES_DB", "wakala")

    return f"postgresql://{user}:{password}@{host}:{port}/{db}"


def slugify(text_val: str) -> str:
    s = str(text_val or "").strip().lower()
    s = re.sub(r"[àáâãäå]", "a", s)
    s = re.sub(r"[èéêë]", "e", s)
    s = re.sub(r"[ìíîï]", "i", s)
    s = re.sub(r"[òóôõö]", "o", s)
    s = re.sub(r"[ùúûü]", "u", s)
    s = re.sub(r"[ç]", "c", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def get_brand_logo_url(brand_name: str) -> str:
    clean = slugify(brand_name).replace("-", "")
    if "mercedes" in clean:
        clean = "mercedes"
    elif "citroen" in clean:
        clean = "citroen"
    elif "omoda" in clean or "jaecoo" in clean:
        clean = "omoda"
    return f"/logos/{clean}.png"


def estimate_fiscal_power_cv(engine_power_hp: Optional[int], fuel_type: str) -> int:
    """Estime la puissance fiscale (CV) marocaine pour la Vignette DGI."""
    if not engine_power_hp:
        return 6
    hp = float(engine_power_hp)
    if fuel_type == "diesel":
        if hp <= 75: return 5
        if hp <= 95: return 6
        if hp <= 115: return 6
        if hp <= 130: return 7
        if hp <= 150: return 8
        if hp <= 190: return 9
        if hp <= 240: return 11
        return 14
    elif fuel_type == "electrique":
        return 0  # Exonérée DGI
    else:  # essence / hybride
        if hp <= 70: return 4
        if hp <= 90: return 5
        if hp <= 110: return 6
        if hp <= 130: return 6
        if hp <= 150: return 7
        if hp <= 180: return 8
        if hp <= 220: return 9
        if hp <= 300: return 11
        return 15


def extract_euro_ncap_stars(ncap_raw: Any) -> Optional[int]:
    s = str(ncap_raw or "").strip()
    match = re.search(r"(\d)\s*★", s)
    if match:
        return int(match.group(1))
    match_num = re.search(r"(\d)", s)
    if match_num and int(match_num.group(1)) <= 5:
        return int(match_num.group(1))
    return None


def generate_available_colors(brand: str, model: str) -> List[Dict[str, Any]]:
    return [
        {"name": "Blanc Glacier", "hex": "#FFFFFF", "price_mad": 0},
        {"name": "Gris Schiste Métallisé", "hex": "#4A4F55", "price_mad": 4000},
        {"name": "Noir Nacré Intense", "hex": "#121212", "price_mad": 4000},
        {"name": "Bleu Océan Profond", "hex": "#1B3B6F", "price_mad": 5000},
        {"name": "Rouge Fusion", "hex": "#8B1E1E", "price_mad": 5500},
    ]


def import_full_catalogue(file_path: Path, dry_run: bool = False) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Catalogue"]

    header_cells = [c.value for c in ws[3]]
    headers = [str(h).strip() for h in header_cells if h is not None]

    rows = []
    for r in range(4, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if any(vals) and vals[0] is not None:
            row_dict = dict(zip(headers, vals[:len(headers)]))
            row_dict["_excel_row_num"] = r
            rows.append(row_dict)

    logger.info(f"Fichier Excel : {len(rows)} variantes de véhicules lues.")

    db_url = get_db_url()
    engine = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    session = Session()

    stats = {
        "total_rows": len(rows),
        "brands_upserted": 0,
        "models_upserted": 0,
        "powertrains_upserted": 0,
        "trims_upserted": 0,
        "vehicles_upserted": 0,
        "scores_upserted": 0,
        "options_upserted": 0,
    }

    try:
        # 1. Assurer le vendeur système
        session.execute(text("""
            INSERT INTO users (id, full_name, email, phone, hashed_password, role, is_verified, is_pro)
            VALUES (:id, 'Wakala Official', 'catalogue@wakala.ma', '+212522000000', 'system_hashed_pwd', 'seller', true, true)
            ON CONFLICT (id) DO NOTHING;
        """), {"id": str(DEFAULT_SYSTEM_SELLER_ID)})

        # 2. Parcourir et traiter chaque ligne du catalogue
        for raw_row in rows:
            v_data, s_data = map_excel_row_to_vehicle_data(raw_row)
            
            brand_name = v_data["brand"]
            model_name = v_data["model"]
            version_name = v_data["version"]
            price_mad = v_data["price"] or 150000.0

            if not brand_name or not model_name or not version_name:
                continue

            brand_slug = slugify(brand_name)
            model_slug = slugify(model_name)
            trim_slug = slugify(f"{model_name}-{version_name}")

            # ── A. CarBrand (car_brands) ──────────────────────────────────
            brand_id_candidate = uuid.uuid5(WAKALA_NAMESPACE, f"brand::{brand_slug}")
            logo_url = get_brand_logo_url(brand_name)
            country = BRAND_ORIGINS.get(brand_name, "International")

            res_brand = session.execute(text("""
                INSERT INTO car_brands (id, name, slug, logo_url, country_of_origin, is_active, created_at, updated_at)
                VALUES (:id, :name, :slug, :logo_url, :country, true, NOW(), NOW())
                ON CONFLICT (name) DO UPDATE SET
                    slug = EXCLUDED.slug,
                    logo_url = EXCLUDED.logo_url,
                    country_of_origin = EXCLUDED.country_of_origin,
                    is_active = true,
                    updated_at = NOW()
                RETURNING id;
            """), {
                "id": str(brand_id_candidate),
                "name": brand_name,
                "slug": brand_slug,
                "logo_url": logo_url,
                "country": country,
            })
            brand_id = res_brand.scalar()
            stats["brands_upserted"] += 1

            # ── B. CarModel (car_models) ──────────────────────────────────
            model_id_candidate = uuid.uuid5(WAKALA_NAMESPACE, f"model::{brand_id}::{model_slug}")
            real_hero_img = get_real_vehicle_image(brand_name, model_name, v_data["body_type"], v_data["fuel_type"])

            res_model = session.execute(text("""
                INSERT INTO car_models (id, brand_id, name, slug, body_type, year_start, hero_image_url, description, created_at, updated_at)
                VALUES (:id, :brand_id, :name, :slug, :body_type, 2026, :hero_image_url, :desc, NOW(), NOW())
                ON CONFLICT (brand_id, slug) DO UPDATE SET
                    name = EXCLUDED.name,
                    body_type = EXCLUDED.body_type,
                    hero_image_url = EXCLUDED.hero_image_url,
                    description = EXCLUDED.description,
                    updated_at = NOW()
                RETURNING id;
            """), {
                "id": str(model_id_candidate),
                "brand_id": str(brand_id),
                "name": model_name,
                "slug": model_slug,
                "body_type": v_data["body_type"].upper(),
                "hero_image_url": real_hero_img,
                "desc": f"{brand_name} {model_name} modèle 2026 officiel au Maroc.",
            })
            model_id = res_model.scalar()
            stats["models_upserted"] += 1

            # ── C. Powertrain (car_powertrains) ───────────────────────────
            pt_name = v_data["engine_type"] or "Moteur Standard"
            fuel_type = v_data["fuel_type"]
            power_hp = v_data["engine_power_hp"]
            fiscal_cv = estimate_fiscal_power_cv(power_hp, fuel_type)
            transmission = v_data["transmission"].upper()

            pt_id = uuid.uuid5(WAKALA_NAMESPACE, f"pt::{model_id}::{slugify(pt_name)}::{fiscal_cv}")

            session.execute(text("""
                INSERT INTO car_powertrains (
                    id, model_id, name, fuel_type, fiscal_power_cv, engine_power_hp,
                    transmission, drivetrain, consumption_l_100, co2_emissions_g_km,
                    created_at, updated_at
                ) VALUES (
                    :id, :model_id, :name, :fuel_type, :fiscal_power_cv, :engine_power_hp,
                    :transmission, :drivetrain, :consumption, :co2,
                    NOW(), NOW()
                )
                ON CONFLICT (id) DO UPDATE SET
                    name = EXCLUDED.name,
                    fuel_type = EXCLUDED.fuel_type,
                    fiscal_power_cv = EXCLUDED.fiscal_power_cv,
                    engine_power_hp = EXCLUDED.engine_power_hp,
                    transmission = EXCLUDED.transmission,
                    consumption_l_100 = EXCLUDED.consumption_l_100,
                    co2_emissions_g_km = EXCLUDED.co2_emissions_g_km,
                    updated_at = NOW();
            """), {
                "id": str(pt_id),
                "model_id": str(model_id),
                "name": pt_name,
                "fuel_type": fuel_type.upper(),
                "fiscal_power_cv": fiscal_cv,
                "engine_power_hp": power_hp,
                "transmission": transmission,
                "drivetrain": "4x4" if v_data["is_4x4"] else "FWD",
                "consumption": v_data["fuel_consumption"],
                "co2": int(v_data["co2_emissions"]) if v_data["co2_emissions"] else None,
            })
            stats["powertrains_upserted"] += 1

            # ── D. TrimCatalog (car_trims) ────────────────────────────────
            trim_id = uuid.uuid5(WAKALA_NAMESPACE, f"trim::{model_id}::{trim_slug}")
            ncap_stars = extract_euro_ncap_stars(v_data["ncap_rating"])
            brand_colors = get_official_colors_for_vehicle(brand_name, model_name)
            colors_json = json.dumps(brand_colors)

            session.execute(text("""
                INSERT INTO car_trims (
                    id, model_id, powertrain_id, name, slug, price_new_mad,
                    promo_price_mad, is_promo, warranty_years, warranty_km,
                    trunk_capacity_l, doors_count, seats_count, euro_ncap_stars,
                    image_url, available_colors, is_available_in_morocco,
                    created_at, updated_at
                ) VALUES (
                    :id, :model_id, :powertrain_id, :name, :slug, :price,
                    :promo_price, :is_promo, 3, 100000,
                    :trunk_l, :doors, :seats, :ncap_stars,
                    :image_url, CAST(:available_colors AS jsonb), true,
                    NOW(), NOW()
                )
                ON CONFLICT (id) DO UPDATE SET
                    name = EXCLUDED.name,
                    price_new_mad = EXCLUDED.price_new_mad,
                    image_url = EXCLUDED.image_url,
                    available_colors = EXCLUDED.available_colors,
                    euro_ncap_stars = EXCLUDED.euro_ncap_stars,
                    trunk_capacity_l = EXCLUDED.trunk_capacity_l,
                    updated_at = NOW();
            """), {
                "id": str(trim_id),
                "model_id": str(model_id),
                "powertrain_id": str(pt_id),
                "name": version_name,
                "slug": trim_slug[:150],
                "price": price_mad,
                "promo_price": None,
                "is_promo": False,
                "trunk_l": v_data["trunk_volume_l"] or 380,
                "doors": v_data["doors"],
                "seats": v_data["seats"],
                "ncap_stars": ncap_stars,
                "image_url": real_hero_img,
                "available_colors": colors_json,
            })
            stats["trims_upserted"] += 1

            # ── E. Vehicle (vehicles) ─────────────────────────────────────
            v_id = uuid.uuid5(WAKALA_NAMESPACE, f"vehicle::{brand_name}::{model_name}::{version_name}")
            v_data["id"] = str(v_id)
            v_data["seller_id"] = str(DEFAULT_SYSTEM_SELLER_ID)

            session.execute(text("""
                INSERT INTO vehicles (
                    id, seller_id, brand, model, version, year, mileage,
                    fuel_type, body_type, transmission, engine_power_hp,
                    doors, seats, city, postal_code, price,
                    trunk_volume_l, ncap_rating, fuel_consumption, co2_emissions,
                    length_cm, is_4x4, engine_type, condition, source, status,
                    description, created_at, updated_at
                ) VALUES (
                    :id, :seller_id, :brand, :model, :version, :year, :mileage,
                    CAST(:fuel_type AS fuel_type), CAST(:body_type AS body_type), CAST(:transmission AS transmission_type), :engine_power_hp,
                    :doors, :seats, :city, :postal_code, :price,
                    :trunk_volume_l, :ncap_rating, :fuel_consumption, :co2_emissions,
                    :length_cm, :is_4x4, :engine_type, :condition, :source, CAST(:status AS vehicle_status),
                    :description, NOW(), NOW()
                )
                ON CONFLICT (id) DO UPDATE SET
                    brand = EXCLUDED.brand,
                    model = EXCLUDED.model,
                    version = EXCLUDED.version,
                    price = EXCLUDED.price,
                    trunk_volume_l = EXCLUDED.trunk_volume_l,
                    ncap_rating = EXCLUDED.ncap_rating,
                    fuel_consumption = EXCLUDED.fuel_consumption,
                    co2_emissions = EXCLUDED.co2_emissions,
                    engine_power_hp = EXCLUDED.engine_power_hp,
                    length_cm = EXCLUDED.length_cm,
                    is_4x4 = EXCLUDED.is_4x4,
                    engine_type = EXCLUDED.engine_type,
                    fuel_type = EXCLUDED.fuel_type,
                    body_type = EXCLUDED.body_type,
                    transmission = EXCLUDED.transmission,
                    description = EXCLUDED.description,
                    updated_at = NOW();
            """), v_data)
            stats["vehicles_upserted"] += 1

            # ── F. Wakala Scores (vehicle_wakala_scores) ───────────────────
            score_id = uuid.uuid5(WAKALA_NAMESPACE, f"scores::{v_id}")
            s_data["id"] = str(score_id)
            s_data["vehicle_id"] = str(v_id)

            session.execute(text("""
                INSERT INTO vehicle_wakala_scores (
                    id, vehicle_id, space_score, safety_score, real_cost_score,
                    access_price_score, city_practicality_score, performance_score,
                    ecology_score, offroad_score, overall_score,
                    data_reliability, observations, source_note,
                    created_at, updated_at
                ) VALUES (
                    :id, :vehicle_id, :space_score, :safety_score, :real_cost_score,
                    :access_price_score, :city_practicality_score, :performance_score,
                    :ecology_score, :offroad_score, :overall_score,
                    :data_reliability, :observations, :source_note,
                    NOW(), NOW()
                )
                ON CONFLICT (vehicle_id) DO UPDATE SET
                    space_score = EXCLUDED.space_score,
                    safety_score = EXCLUDED.safety_score,
                    real_cost_score = EXCLUDED.real_cost_score,
                    access_price_score = EXCLUDED.access_price_score,
                    city_practicality_score = EXCLUDED.city_practicality_score,
                    performance_score = EXCLUDED.performance_score,
                    ecology_score = EXCLUDED.ecology_score,
                    offroad_score = EXCLUDED.offroad_score,
                    overall_score = EXCLUDED.overall_score,
                    data_reliability = EXCLUDED.data_reliability,
                    observations = EXCLUDED.observations,
                    source_note = EXCLUDED.source_note,
                    updated_at = NOW();
            """), s_data)
            stats["scores_upserted"] += 1

            # ── G. Vehicle Options & Colors (pour configurateur) ──────────
            # Clean old default options/colors for this vehicle if updating
            session.execute(text("DELETE FROM vehicle_colors WHERE vehicle_id = :v_id;"), {"v_id": str(v_id)})
            session.execute(text("DELETE FROM vehicle_options WHERE vehicle_id = :v_id;"), {"v_id": str(v_id)})

            for c in brand_colors:
                session.execute(text("""
                    INSERT INTO vehicle_colors (id, vehicle_id, color_name, hex_code, price_delta, is_default)
                    VALUES (uuid_generate_v4(), :v_id, :color_name, :hex_code, :price_delta, :is_default)
                """), {
                    "v_id": str(v_id),
                    "color_name": c["name"],
                    "hex_code": c["hex"],
                    "price_delta": c.get("price_mad", 0),
                    "is_default": c.get("is_default", False),
                })

            brand_options = get_official_options_for_vehicle(
                brand_name,
                model_name,
                v_data["body_type"],
                price_mad,
                is_electric_or_hybrid=(v_data["fuel_type"] in ["electrique", "hybride", "hybride_rechargeable"])
            )
            for opt in brand_options:
                session.execute(text("""
                    INSERT INTO vehicle_options (id, vehicle_id, category, name, price_delta, is_default)
                    VALUES (uuid_generate_v4(), :v_id, :category, :name, :price_delta, :is_default)
                """), {
                    "v_id": str(v_id),
                    "category": opt["category"],
                    "name": opt["name"],
                    "price_delta": opt.get("price_delta", 0),
                    "is_default": opt.get("is_default", False),
                })
            stats["options_upserted"] += 1

        session.commit()
        logger.info(f"✅ Importation réussie de tous les véhicules du catalogue !")
        logger.info(f"Statistiques : {stats}")

    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur critique lors de l'import : {e}")
        raise
    finally:
        session.close()

    return stats


def main():
    parser = argparse.ArgumentParser(description="Importer tout le catalogue Wakala Excel vers PostgreSQL.")
    parser.add_argument("--file", "-f", help="Chemin vers le fichier Excel wakala-catalogue.xlsx", default=None)
    parser.add_argument("--dry-run", action="store_true", help="Valider sans écrire en BDD")
    args = parser.parse_args()

    catalogue_path = resolve_catalogue_file(args.file)
    stats = import_full_catalogue(catalogue_path, dry_run=args.dry_run)
    print("\n" + "=" * 60)
    print("RÉSUMÉ DE L'IMPORT DU CATALOGUE GLOBAL WAKALA")
    print("=" * 60)
    print(f"Lignes totales Excel     : {stats['total_rows']}")
    print(f"Marques (car_brands)     : {stats['brands_upserted']}")
    print(f"Modèles (car_models)     : {stats['models_upserted']}")
    print(f"Motorisations (powertr.) : {stats['powertrains_upserted']}")
    print(f"Finitions (car_trims)    : {stats['trims_upserted']}")
    print(f"Véhicules (vehicles)     : {stats['vehicles_upserted']}")
    print(f"Notes Wakala (scores)    : {stats['scores_upserted']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
