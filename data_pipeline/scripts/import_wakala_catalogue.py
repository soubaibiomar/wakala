#!/usr/bin/env python3
"""
import_wakala_catalogue.py — Importe le catalogue officiel Wakala (Excel) dans PostgreSQL.
========================================================================================

Lit le fichier wakala-catalogue.xlsx (feuille "Catalogue"), valide chaque ligne via SchemaValidator,
génère un vehicle_id stable et déterministe (UUIDv5) et insère les données dans les tables
'vehicles' et 'vehicle_wakala_scores' de manière 100% idempotente.

Usage :
    python -m data_pipeline.scripts.import_wakala_catalogue
    python -m data_pipeline.scripts.import_wakala_catalogue --file /path/to/wakala-catalogue.xlsx
    python -m data_pipeline.scripts.import_wakala_catalogue --dry-run
"""

import argparse
import logging
import os
import sys
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Configuration du PYTHONPATH
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "data_pipeline" / "kafka" / "producers" / "scrapers"))

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from data_pipeline.kafka.producers.scrapers.schema_validator import SchemaValidator
from data_pipeline.scripts.catalogue_mapping import (
    map_excel_row_to_vehicle_data,
    EXCEL_COLUMN_HEADERS,
)

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("WakalaCatalogueImporter")

# UUID Namespace fixe pour la génération déterministe des UUIDs
WAKALA_NAMESPACE = uuid.UUID("e743a18e-42c2-4876-9051-b841e4eb4192")
DEFAULT_SYSTEM_SELLER_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")

# Chemins de recherche automatique du fichier catalogue
CANDIDATE_PATHS = [
    Path(r"C:\Users\omar\Downloads\wakala-catalogue.xlsx"),
    PROJECT_ROOT.parent / "wakala-catalogue-final.xlsx",
    PROJECT_ROOT.parent / "wakala-cataloguel.xlsx",
    PROJECT_ROOT / "wakala-catalogue.xlsx",
]


def resolve_catalogue_file(custom_path: Optional[str] = None) -> Path:
    """Trouve le fichier catalogue Excel valide."""
    if custom_path:
        p = Path(custom_path)
        if p.exists():
            return p
        raise FileNotFoundError(f"Fichier catalogue non trouvé à l'emplacement : {custom_path}")

    for p in CANDIDATE_PATHS:
        if p.exists():
            logger.info(f"Fichier catalogue détecté : {p}")
            return p

    raise FileNotFoundError(
        "Fichier wakala-catalogue.xlsx introuvable. "
        "Veuillez spécifier le chemin via l'option --file."
    )


def generate_deterministic_vehicle_id(brand: str, model: str, version: str) -> uuid.UUID:
    """
    Génère un UUIDv5 déterministe basé sur marque + modèle + version normalisés.
    Idempotent : deux exécutions avec les mêmes entrées produisent strictement le même UUID.
    """
    key = f"{brand.lower().strip()}::{model.lower().strip()}::{version.lower().strip()}"
    return uuid.uuid5(WAKALA_NAMESPACE, key)


def get_db_url() -> str:
    """Construit l'URL de connexion PostgreSQL depuis les variables d'environnement."""
    # Chargement direct du fichier .env si présent
    env_file = PROJECT_ROOT / ".env"
    if env_file.exists():
        with open(env_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    if k not in os.environ:
                        os.environ[k] = v.strip('"').strip("'")

    db_url = os.getenv("DATABASE_URL")
    if db_url:
        return db_url.replace("postgresql+asyncpg://", "postgresql://")

    user = os.getenv("POSTGRES_USER", "wakala_user")
    password = os.getenv("POSTGRES_PASSWORD", "wakala_secret_password")
    host = os.getenv("POSTGRES_HOST", "localhost")
    port = os.getenv("POSTGRES_PORT", "5433")
    db = os.getenv("POSTGRES_DB", "wakala")

    return f"postgresql://{user}:{password}@{host}:{port}/{db}"


def load_catalogue_rows(file_path: Path) -> List[Dict[str, Any]]:
    """Lit la feuille 'Catalogue' du fichier Excel à partir de la ligne d'en-tête (ligne 3)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Catalogue" not in wb.sheetnames:
        raise ValueError(f"Feuille 'Catalogue' manquante dans {file_path}. Feuilles: {wb.sheetnames}")

    ws = wb["Catalogue"]
    header_cells = [c.value for c in ws[3]]
    headers = [str(h).strip() for h in header_cells if h is not None]

    rows = []
    for r in range(4, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        # Si la ligne contient au moins la marque ou un modèle
        if any(vals) and vals[0] is not None:
            row_dict = dict(zip(headers, vals[:len(headers)]))
            row_dict["_excel_row_num"] = r
            rows.append(row_dict)

    logger.info(f"Fichier Excel chargé avec succès : {len(rows)} lignes lues depuis la feuille 'Catalogue'.")
    return rows


def validate_and_transform_row(
    raw_row: Dict[str, Any]
) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]], List[str]]:
    """
    Valide et transforme une ligne brute en enregistrements pour 'vehicles' et 'vehicle_wakala_scores'.
    Retourne (vehicle_data, wakala_scores_data, errors).
    """
    errors = []
    row_num = raw_row.get("_excel_row_num", "?")

    # 1. Extraction et mapping
    try:
        vehicle_data, wakala_scores = map_excel_row_to_vehicle_data(raw_row)
    except Exception as e:
        errors.append(f"Erreur de transformation ligne {row_num}: {e}")
        return None, None, errors

    # 2. Validation champs obligatoires de base
    brand = vehicle_data.get("brand")
    model = vehicle_data.get("model")
    version = vehicle_data.get("version")
    price = vehicle_data.get("price")

    if not brand or brand == "unknown" or brand == "N/A":
        errors.append(f"Ligne {row_num}: Champ obligatoire manquant 'Marque'")
    if not model or model == "unknown":
        errors.append(f"Ligne {row_num}: Champ obligatoire manquant 'Modèle'")
    if not version:
        errors.append(f"Ligne {row_num}: Champ obligatoire manquant 'Variante'")
    if price is None:
        errors.append(f"Ligne {row_num}: Prix manquant ou invalide")

    # 3. Validation via SchemaValidator
    validator_payload = {
        "brand": brand,
        "source": vehicle_data.get("source", "wakala_catalogue"),
        "source_url": f"https://wakala.ma/catalogue/{brand}/{model}",
        "price": price,
        "year": vehicle_data.get("year", 2026),
        "mileage": vehicle_data.get("mileage", 0),
    }
    is_valid, val_errors = SchemaValidator.validate(validator_payload)
    if not is_valid:
        for err in val_errors:
            errors.append(f"Ligne {row_num} [SchemaValidator]: {err}")

    if errors:
        return None, None, errors

    # 4. Génération de l'ID déterministe
    v_id = generate_deterministic_vehicle_id(brand, model, version)
    vehicle_data["id"] = str(v_id)
    vehicle_data["seller_id"] = str(DEFAULT_SYSTEM_SELLER_ID)

    wakala_scores["id"] = str(uuid.uuid5(WAKALA_NAMESPACE, f"scores::{v_id}"))
    wakala_scores["vehicle_id"] = str(v_id)

    return vehicle_data, wakala_scores, []


def ensure_system_seller(session):
    """S'assure que l'utilisateur vendeur système officiel Wakala existe."""
    insert_seller_sql = text("""
        INSERT INTO users (id, name, email, password_hash, role, is_verified, is_pro)
        VALUES (:id, 'Wakala Official', 'catalogue@wakala.ma', 'system_hashed_pwd', 'seller', true, true)
        ON CONFLICT (id) DO NOTHING;
    """)
    session.execute(insert_seller_sql, {"id": str(DEFAULT_SYSTEM_SELLER_ID)})


def import_catalogue(file_path: Path, dry_run: bool = False) -> Dict[str, Any]:
    """Exécute l'import complet du catalogue Excel."""
    raw_rows = load_catalogue_rows(file_path)

    valid_vehicles: List[Dict[str, Any]] = []
    valid_scores: List[Dict[str, Any]] = []
    rejected_rows: List[Dict[str, Any]] = []

    for row in raw_rows:
        v_data, s_data, errors = validate_and_transform_row(row)
        if errors:
            rejected_rows.append({
                "row_num": row.get("_excel_row_num"),
                "brand": row.get(EXCEL_COLUMN_HEADERS["brand"]),
                "model": row.get(EXCEL_COLUMN_HEADERS["model"]),
                "version": row.get(EXCEL_COLUMN_HEADERS["version"]),
                "errors": errors,
            })
            for err in errors:
                logger.warning(f"Rejet ligne: {err}")
        else:
            valid_vehicles.append(v_data)
            valid_scores.append(s_data)

    stats = {
        "total_read": len(raw_rows),
        "valid_count": len(valid_vehicles),
        "rejected_count": len(rejected_rows),
        "inserted_or_updated": 0,
    }

    logger.info(
        f"Validation terminée : {stats['valid_count']} véhicules valides, "
        f"{stats['rejected_count']} lignes rejetées."
    )

    if dry_run:
        logger.info("[DRY-RUN] Aucune écriture en base de données.")
        return stats

    # Connexion et insertion en base
    db_url = get_db_url()
    engine = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    session = Session()

    try:
        # 1. Vérifier/Créer le vendeur système
        ensure_system_seller(session)

        # 2. Insertion idempotente dans 'vehicles'
        upsert_vehicle_sql = text("""
            INSERT INTO vehicles (
                id, seller_id, brand, model, version, year, mileage,
                fuel_type, body_type, transmission, engine_power_hp,
                doors, seats, city, postal_code, price,
                trunk_volume_l, ncap_rating, fuel_consumption, co2_emissions,
                length_cm, is_4x4, engine_type, condition, source, status,
                description, created_at, updated_at
            ) VALUES (
                :id, :seller_id, :brand, :model, :version, :year, :mileage,
                CAST(:fuel_type AS fuel_type), CAST(:body_type AS body_type), CAST(:transmission AS transmission_type), :engine_power_hp,
                :doors, :seats, :city, :postal_code, :price,
                :trunk_volume_l, :ncap_rating, :fuel_consumption, :co2_emissions,
                :length_cm, :is_4x4, :engine_type, :condition, :source, CAST(:status AS vehicle_status),
                :description, NOW(), NOW()
            )
            ON CONFLICT (id) DO UPDATE SET
                seller_id = EXCLUDED.seller_id,
                brand = EXCLUDED.brand,
                model = EXCLUDED.model,
                version = EXCLUDED.version,
                price = EXCLUDED.price,
                trunk_volume_l = EXCLUDED.trunk_volume_l,
                ncap_rating = EXCLUDED.ncap_rating,
                fuel_consumption = EXCLUDED.fuel_consumption,
                co2_emissions = EXCLUDED.co2_emissions,
                engine_power_hp = EXCLUDED.engine_power_hp,
                length_cm = EXCLUDED.length_cm,
                is_4x4 = EXCLUDED.is_4x4,
                engine_type = EXCLUDED.engine_type,
                fuel_type = EXCLUDED.fuel_type,
                body_type = EXCLUDED.body_type,
                transmission = EXCLUDED.transmission,
                doors = EXCLUDED.doors,
                seats = EXCLUDED.seats,
                condition = EXCLUDED.condition,
                source = EXCLUDED.source,
                description = EXCLUDED.description,
                updated_at = NOW();
        """)

        # 3. Insertion idempotente dans 'vehicle_wakala_scores'
        upsert_scores_sql = text("""
            INSERT INTO vehicle_wakala_scores (
                id, vehicle_id, space_score, safety_score, real_cost_score,
                access_price_score, city_practicality_score, performance_score,
                ecology_score, offroad_score, overall_score,
                data_reliability, observations, source_note,
                created_at, updated_at
            ) VALUES (
                :id, :vehicle_id, :space_score, :safety_score, :real_cost_score,
                :access_price_score, :city_practicality_score, :performance_score,
                :ecology_score, :offroad_score, :overall_score,
                :data_reliability, :observations, :source_note,
                NOW(), NOW()
            )
            ON CONFLICT (vehicle_id) DO UPDATE SET
                space_score = EXCLUDED.space_score,
                safety_score = EXCLUDED.safety_score,
                real_cost_score = EXCLUDED.real_cost_score,
                access_price_score = EXCLUDED.access_price_score,
                city_practicality_score = EXCLUDED.city_practicality_score,
                performance_score = EXCLUDED.performance_score,
                ecology_score = EXCLUDED.ecology_score,
                offroad_score = EXCLUDED.offroad_score,
                overall_score = EXCLUDED.overall_score,
                data_reliability = EXCLUDED.data_reliability,
                observations = EXCLUDED.observations,
                source_note = EXCLUDED.source_note,
                updated_at = NOW();
        """)

        # Exécution par lots
        for v in valid_vehicles:
            session.execute(upsert_vehicle_sql, v)

        for s in valid_scores:
            session.execute(upsert_scores_sql, s)

        session.commit()
        stats["inserted_or_updated"] = len(valid_vehicles)
        logger.info(f"✅ Insertion réussie : {stats['inserted_or_updated']} véhicules & scores insérés/mis à jour en base.")

    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur lors de l'insertion en base : {e}")
        raise
    finally:
        session.close()

    return stats


def main():
    parser = argparse.ArgumentParser(description="Importer le catalogue officiel Wakala (Excel) vers PostgreSQL.")
    parser.add_argument("--file", "-f", help="Chemin vers le fichier Excel wakala-catalogue.xlsx", default=None)
    parser.add_argument("--dry-run", action="store_true", help="Valider et parser sans écrire en base de données")
    args = parser.parse_args()

    try:
        catalogue_path = resolve_catalogue_file(args.file)
        stats = import_catalogue(catalogue_path, dry_run=args.dry_run)
        print("\n" + "=" * 50)
        print("RÉSUMÉ DE L'IMPORT DU CATALOGUE WAKALA")
        print("=" * 50)
        print(f"Lignes totales lues    : {stats['total_read']}")
        print(f"Véhicules valides      : {stats['valid_count']}")
        print(f"Lignes rejetées        : {stats['rejected_count']}")
        print(f"Enregistrés en BDD     : {stats['inserted_or_updated']}")
        print("=" * 50)
    except Exception as e:
        logger.critical(f"Échec critique de l'import : {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
